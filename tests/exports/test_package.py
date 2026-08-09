"""Export ZIP + manifest: employee-id filenames, checksum, expiry, round-trip."""

from __future__ import annotations

import datetime as dt
import io
import zipfile

import pytest
from openpyxl import load_workbook

from apps.exports.models import ExportManifest
from apps.exports.services import create_export
from apps.payroll.services.close import close_period, prepare_period

pytestmark = pytest.mark.django_db

PAYLOAD = {
    "pay_date": "2026-08-05",
    "calc_period": "2026-07",
    "lines": [
        {
            "employee_id": "EMP-0001",
            "department": "베이커리",
            "title": "사원",
            "net": 1_904_000,
            "insurance_final": True,
            "earnings": [
                {"label": "기본급", "amount": 1_800_000},
                {"label": "주휴수당", "amount": 200_000},
            ],
            "deductions": [
                {"label": "국민연금", "amount": 81_000},
                {"label": "소득세", "amount": 15_000},
            ],
            "detail_lines": ["기본급: 160.00h × 11,250"],
        },
        {
            "employee_id": "EMP-0002",
            "department": "매장",
            "title": "매니저",
            "net": 2_500_000,
            "insurance_final": True,
            "earnings": [{"label": "기본급", "amount": 2_600_000}],
            "deductions": [{"label": "국민연금", "amount": 100_000}],
            "detail_lines": ["기본급: 월 고정"],
        },
    ],
}


def _snapshot():
    period = prepare_period(dt.date(2026, 7, 1))
    return close_period(period=period, payload=PAYLOAD)


def test_zip_uses_employee_id_filenames_and_manifest() -> None:
    snap = _snapshot()
    manifest_obj, zip_bytes = create_export(snapshot=snap, requester=None)

    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = zf.namelist()
        assert "manifest.json" in names
        assert any(n.startswith("summary-2026-07") for n in names)
        assert "pay-statement-EMP-0001-2026-07.xlsx" in names
        assert "pay-statement-EMP-0002-2026-07.xlsx" in names
        # No employee name ever appears in a filename.
        assert not any("강혜령" in n or "현희" in n for n in names)

        # Reverse-verify a statement's net against the payload via openpyxl.
        with zf.open("pay-statement-EMP-0001-2026-07.xlsx") as fh:
            ws = load_workbook(io.BytesIO(fh.read())).active
            assert ws is not None
            assert ws["D25"].value == 1_904_000

    assert manifest_obj.checksum == snap.checksum
    assert manifest_obj.manifest["snapshot_version"] == snap.version


def test_manifest_download_expires_in_five_minutes() -> None:
    snap = _snapshot()
    manifest_obj, _ = create_export(snapshot=snap, requester=None)
    # Fresh manifest is live; a past expiry is treated as expired (private URL gate).
    assert manifest_obj.is_expired is False
    delta = manifest_obj.expires_at - manifest_obj.created_at
    assert dt.timedelta(minutes=4) <= delta <= dt.timedelta(minutes=6)

    manifest_obj.expires_at = manifest_obj.created_at - dt.timedelta(seconds=1)
    ExportManifest.objects.filter(pk=manifest_obj.pk).update(expires_at=manifest_obj.expires_at)
    manifest_obj.refresh_from_db()
    assert manifest_obj.is_expired is True

"""Statement workbook: totals-as-values, integer-only, formula-free, PII-free."""

from __future__ import annotations

import datetime as dt

from openpyxl.worksheet.worksheet import Worksheet

from apps.exports.statement import LineItem, StatementData, build_statement

FORBIDDEN = ["강혜령", "현희", "TODAY", "=B8", "0.2"]


def _ws() -> Worksheet:
    ws = build_statement(_data()).active
    assert ws is not None
    return ws


def _data() -> StatementData:
    return StatementData(
        attribution_month=dt.date(2026, 7, 1),
        employee_id="EMP-0001",
        department="베이커리",
        title="사원",
        pay_date=dt.date(2026, 8, 5),
        calc_period="2026-07",
        earnings=[LineItem("기본급", 1_800_000), LineItem("주휴수당", 200_000)],
        deductions=[LineItem("국민연금", 81_000), LineItem("소득세", 15_000)],
        detail_lines=["기본급: 160.00h × 11,250 = 1,800,000"],
        version=1,
        checksum="abc123checksum",
    )


def test_single_visible_sheet() -> None:
    wb = build_statement(_data())
    assert len(wb.worksheets) == 1
    assert wb.worksheets[0].sheet_state == "visible"


def test_totals_are_static_values_not_formulas() -> None:
    ws = _ws()
    b_sum = sum(ws[f"B{r}"].value or 0 for r in range(8, 24))
    d_sum = sum(ws[f"D{r}"].value or 0 for r in range(8, 24))
    assert ws["B25"].value == b_sum == 2_000_000
    assert ws["D24"].value == d_sum == 96_000
    assert ws["D25"].value == ws["B25"].value - ws["D24"].value == 1_904_000
    for coord in ("B25", "D24", "D25"):
        assert ws[coord].data_type != "f"


def test_all_money_cells_are_integers() -> None:
    ws = _ws()
    for coord in ("B8", "B9", "D8", "D9", "B25", "D24", "D25"):
        assert isinstance(ws[coord].value, int)


def test_no_formulas_and_no_pii_anywhere() -> None:
    ws = _ws()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                assert not cell.value.startswith("=")
                for bad in FORBIDDEN:
                    assert bad not in cell.value
            assert cell.data_type != "f"


def test_month_paydate_and_checksum_present() -> None:
    ws = _ws()
    assert ws["B1"].value == "2026-07"
    texts = [c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    assert any("지급일: 2026-08-05" in t for t in texts)
    assert any("abc123checksum" in t for t in texts)
    assert any("산정 내역" in t for t in texts)


def test_a4_portrait_one_page_wide() -> None:
    ws = _ws()
    assert ws.page_setup.orientation == "portrait"
    assert ws.page_setup.paperSize == 9
    assert ws.page_setup.fitToWidth == 1
    page_props = ws.sheet_properties.pageSetUpPr
    assert page_props is not None
    assert page_props.fitToPage is True

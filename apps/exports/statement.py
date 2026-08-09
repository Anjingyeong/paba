"""Build a single-sheet, PII-free, formula-free pay statement workbook.

The layout follows the original workbooks' *visual* structure (a reference only —
the originals are never read, registered, or shipped) while removing all personal
information and the erroneous/volatile formulas. Every monetary cell is a static
integer computed in Python; the sheet contains no worksheet formulas at all, so
there is nothing to recalculate and no ``TODAY()``/``B8*0.2`` to go wrong.

Cell map (per the plan):
- ``B1``      attribution month (YYYY-MM)
- ``A3:B4``   employee identity (opaque id) and department/title
- ``A8:B23``  earnings labels/amounts
- ``C8:D23``  deduction labels/amounts
- ``D24``     total deductions
- ``B25``     total pay      ``D25`` net pay
- row 27+     legal detail lines (pay date, period, formulas, weekly basis,
              issue version + snapshot checksum)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet

FIRST_ROW = 8
LAST_ITEM_ROW = 23  # 16 item rows (8..23)
MONEY_FORMAT = "#,##0"


@dataclass(frozen=True)
class LineItem:
    label: str
    amount: int


@dataclass
class StatementData:
    attribution_month: date
    employee_id: str
    department: str
    title: str
    pay_date: date
    calc_period: str
    earnings: list[LineItem] = field(default_factory=list)
    deductions: list[LineItem] = field(default_factory=list)
    detail_lines: list[str] = field(default_factory=list)
    version: int = 1
    checksum: str = ""

    @property
    def total_pay(self) -> int:
        return sum(i.amount for i in self.earnings)

    @property
    def total_deduction(self) -> int:
        return sum(i.amount for i in self.deductions)

    @property
    def net_pay(self) -> int:
        return self.total_pay - self.total_deduction


def _money(ws: Worksheet, coord: str, value: int) -> None:
    cell = ws[coord]
    cell.value = value
    cell.number_format = MONEY_FORMAT


def build_statement(data: StatementData) -> Workbook:
    if len(data.earnings) > (LAST_ITEM_ROW - FIRST_ROW + 1):
        raise ValueError("Too many earnings rows for the statement layout.")
    if len(data.deductions) > (LAST_ITEM_ROW - FIRST_ROW + 1):
        raise ValueError("Too many deduction rows for the statement layout.")

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "급여명세서"

    ws["A1"] = "급여명세서"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "귀속월"
    ws["B1"] = f"{data.attribution_month:%Y-%m}"

    ws["A3"] = "직원식별"
    ws["B3"] = data.employee_id
    ws["A4"] = "부서/직급"
    ws["B4"] = f"{data.department} / {data.title}"

    ws["A7"] = "지급 항목"
    ws["B7"] = "금액"
    ws["C7"] = "공제 항목"
    ws["D7"] = "금액"
    for cell in ("A7", "B7", "C7", "D7"):
        ws[cell].font = Font(bold=True)

    for offset, item in enumerate(data.earnings):
        row = FIRST_ROW + offset
        ws[f"A{row}"] = item.label
        _money(ws, f"B{row}", item.amount)
    for offset, item in enumerate(data.deductions):
        row = FIRST_ROW + offset
        ws[f"C{row}"] = item.label
        _money(ws, f"D{row}", item.amount)

    # Totals as static integers (never worksheet formulas).
    ws["C24"] = "총공제"
    _money(ws, "D24", data.total_deduction)
    ws["A25"] = "총지급"
    _money(ws, "B25", data.total_pay)
    ws["C25"] = "실지급액"
    _money(ws, "D25", data.net_pay)
    for cell in ("A25", "C25"):
        ws[cell].font = Font(bold=True)

    # Legal detail block.
    detail_start = 27
    header = ws[f"A{detail_start}"]
    header.value = "산정 내역"
    header.font = Font(bold=True)
    lines = [
        f"지급일: {data.pay_date:%Y-%m-%d}",
        f"산정기간: {data.calc_period}",
        *data.detail_lines,
        f"발급버전: v{data.version}",
        f"스냅샷 체크섬: {data.checksum}",
    ]
    last_row = detail_start
    for i, line in enumerate(lines, start=detail_start + 1):
        ws[f"A{i}"] = line
        last_row = i

    # Column widths and one-page-wide A4 portrait print setup.
    for col, width in (("A", 22), ("B", 16), ("C", 22), ("D", 16)):
        ws.column_dimensions[col].width = width
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_area = f"A1:D{last_row}"

    return wb

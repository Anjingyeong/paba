"""Manager summary workbook: one row per employee with totals (static integers)."""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font

from .statement import MONEY_FORMAT, StatementData


def build_summary(month_label: str, statements: list[StatementData]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "요약"

    ws["A1"] = f"{month_label} 급여 요약"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["직원식별", "총지급", "총공제", "실지급"]
    for col, text in zip("ABCD", headers, strict=True):
        cell = ws[f"{col}3"]
        cell.value = text
        cell.font = Font(bold=True)

    row = 4
    total_pay = total_ded = total_net = 0
    for data in statements:
        ws[f"A{row}"] = data.employee_id
        for col, value in (("B", data.total_pay), ("C", data.total_deduction), ("D", data.net_pay)):
            ws[f"{col}{row}"].value = value
            ws[f"{col}{row}"].number_format = MONEY_FORMAT
        total_pay += data.total_pay
        total_ded += data.total_deduction
        total_net += data.net_pay
        row += 1

    ws[f"A{row}"] = "합계"
    ws[f"A{row}"].font = Font(bold=True)
    for col, value in (("B", total_pay), ("C", total_ded), ("D", total_net)):
        ws[f"{col}{row}"].value = value
        ws[f"{col}{row}"].number_format = MONEY_FORMAT
        ws[f"{col}{row}"].font = Font(bold=True)

    for col, width in (("A", 22), ("B", 16), ("C", 16), ("D", 16)):
        ws.column_dimensions[col].width = width
    return wb
